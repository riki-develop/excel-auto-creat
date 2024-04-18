import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

def create_excel_file():
    # 今日の日付を取得
    today = datetime.date.today()

    # 土日の場合は処理をスキップ
    if today.weekday() >= 5:  # 5 = Saturday, 6 = Sunday
        return

    # 年月のフォルダを作成
    month_folder = today.strftime('%Y%m')
    base_path = "/Users/chikaras/Works/1.publicis/documents/_日報実績シート/"
    month_path = os.path.join(base_path, month_folder)
    os.makedirs(month_path, exist_ok=True)

    # ファイル名を設定
    file_name = f"WFH Daily Report ‐ Chikara Karasawa - {today.strftime('%Y%m%d')}.xlsx"
    file_path = os.path.join(month_path, file_name)

    # 直近のファイルを探す
    try:
        files = [f for f in os.listdir(month_path) if f.endswith('.xlsx')]
        latest_file = max([os.path.join(month_path, f) for f in files], key=os.path.getctime)
        wb = load_workbook(latest_file)
    except (FileNotFoundError, ValueError):
        # 月の初日またはファイルが見つからない場合は新しいワークブックを作成
        wb = Workbook()

    ws = wb.active
    ws['C3'] = today.strftime('%Y/%m/%d')  # エクセル内表示用日付
    ws['C3'].alignment = Alignment(horizontal='right')  # 右寄せに設定

    # ファイルを保存
    wb.save(file_path)

if __name__ == "__main__":
    create_excel_file()